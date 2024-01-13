import os
import pandas as pd
from openpyxl import Workbook, load_workbook
import csv

filename = "results.xlsx"
wb = Workbook()
sheet = wb.active
sheet["A1"] = "hello"
sheet["B1"] = "World"
wb.save(filename=filename)

print(filename)




data = ["Swag",
          "SwagCase",
          "numberSwag"
          "DeadSwag",
          "ICantSPellPetitionerTosavemylife",
          "AttorneyKnowMySwagNotMyStory"
          ]
case_dict = {"filing_date":[i for i in data if "Swag" in i],
                   "case_number":[i for i in data if "numberSwag" in i],
                   "decedntInfo":[i for i in data if "DeadSwag" in i],
                   "petitionerInfo":[i for i in data if "ICantSPellPetitionerTosavemylife" in i],
                   "attorneyInfo":[i for i in data if "AttorneyKnowMySwagNotMyStory" in i]}

# u = create_wb()
# # with open(filepath, "a", newline='') as csvfile:
# #       writer = csv.DictWriter(csvfile, fieldnames=case_dict.keys())
# #       if csvfile.tell() == 0:
# #             writer.writeheader()

# #             for i in range(len(case_dict['filing_date'])):
# #                   row = {key: case_dict[i] for key in case_dict.keys()}
# #                   writer.writerow(row)
# #                   writer.save()




# def xcel_check(self, file_path):
#       # check if workbok is on the machine
#       file_path = "C:/Users/P3128232/florida_boyz/results.xslx"
#       if os.isfile(file_path):
#             if file_path.endswith(".xslx"):
#                   # get_excel file function
#                   print ("File exists")
#             return True
#       else:
#             pass


# # def get_excel_file(self, file_path):
# #       wb = load_workbook(filename="results.xlsx")
# #       pass

# # def create_wb():
# #       wb = Workbook('results.xlsx')
# #       ws = wb.active
# #       ws.title = "Probate Cases"
# #       columns = ['Filing Date', 'Case Number', 'Decedent info', 'Petitioner info', 'Attorney info']
# #       for i , name in enumerate(columns, start=1):
# #             ws.cell(row=1, column=i).value= name


# def clean_data(self, data):
#       data = ["Swag",
#           "SwagCase",
#           "numberSwag"
#           "DeadSwag",
#           "ICantSPellPetitionerTosavemylife",
#           "AttorneyKnowMySwagNotMyStory"
#           ]
#       case_dict = {"filing_date":[i for i in data if "Swag" in i],
#                    "case_number":[i for i in data if "numberSwag" in i],
#                    "decedntInfo":[i for i in data if "DeadSwag" in i],
#                    "petitionerInfo":[i for i in data if "ICantSPellPetitionerTosavemylife" in i],
#                    "attorneyInfo":[i for i in data if "AttorneyKnowMySwagNotMyStory" in i]}
#       return case_dict
